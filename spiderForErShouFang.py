import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
import time
class spiderForErShouFang:
    '''
    initialize class,get the value of cityList, and set value for linkDict and cityDict
    Args:
        cityList: the list of cities we intend to crawl from Lianjia, set the default value be empty
    '''
    def __init__(self,cityList,cityDict):
        self.cityList = cityList
        self.linkDict = {}
        self.cityDict= cityDict
        self.linkNewDict={}
    '''
    get links of all the regions(in detail),both save the results in the linkDict(the attr of the class) and files in the ./doc/link named after the city

    '''
    def getLinks(self):
        cityList =self.cityList

        for city in cityList:
            linkList = []
            link_all = "https://"+city+".lianjia.com"
            link_one = "https://"+city+".lianjia.com/ershoufang/"
            response_one = requests.get(link_one,verify=False)
            soup_one = BeautifulSoup(response_one.text,"lxml")
            divErShouFang = soup_one.find_all("div",attrs={"data-role":"ershoufang"})[0]#根据html5元素分析页面
            regionTwoListLinks = divErShouFang.find_all("a")
            for regionTwoLink in regionTwoListLinks:
                remotePlace = False
                if regionTwoLink["href"].startswith("https"):
                    link_two = regionTwoLink["href"]
                else:
                    link_two = link_all + regionTwoLink["href"]
                if not link_two.startswith("https//"+city):
                    remotePlace = True
                response_two =requests.get(link_two,verify=False)
                soup_two = BeautifulSoup(response_two.text,"lxml")
                divErShouFang2 = soup_two.find_all("div",attrs={"data-role":"ershoufang"})[0].find_all("div")[1]

                regionThreeListLinks = divErShouFang2.find_all("a")
                for regionThreeLink in regionThreeListLinks:
                    if remotePlace:
                        link_three = "https://"+link_two.split("/")[2]+regionThreeLink["href"]
                    else:
                        link_three = link_all + regionThreeLink["href"]
                    if not  link_three in linkList:
                        linkList.append(link_three)
                        with open("./doc/link/ershoufang-"+city+".txt","a+",encoding="utf-8") as f:
                            f.write(link_three+"\n")
            self.linkDict[city] = linkList
    '''
        judge whether we have gotten the links where the specific information of housing estate comes from. If not,
        then try to get the links from the files saved in the .doc/link
        Args:
            city: the city we will check whether we have gotten the link list
    '''
    def ensureLinks(self,city):
        if not city in self.linkDict.keys() or len(self.linkDict[city])<1:
            print(True)
            with open("./doc/link/ershoufang-"+city+".txt","r",encoding="utf-8") as f:
                linkList = []
                lines = list(f)
                for line in lines:
                    linkList.append(line.split("\n")[0])
                self.linkDict[city] = linkList
                print("ensureLinks:",self.linkDict[city])
    '''
        crawl the page with specific information of housing estate,and save data in ./doc/data.xlsx
        Args:
            link: the link of the page we intend to crawl
            city: the city we intend to crawl
    '''
    def getDetailPage(self,link,city):
        print("link",link)
        cityDict=self.cityDict
        regionOne = self.cityDict[city]
        print(regionOne)
        count = 0
        others = False
        cityFromURL= link.split("//")[1].split(".")[0]
        others = (city!=cityFromURL)
        pageLinks=[]
        while count <5:
            try:
                responseOne = requests.get(link)
                soup_one =    BeautifulSoup(responseOne.text,"lxml")
                numOfPage = eval(soup_one.find_all("div",attrs={"class":"page-box house-lst-page-box"})[0]["page-data"])["totalPage"]
                regionTwo = soup_one.find_all("a",attrs={"class":"selected"})[1].text.strip()
                regionThree = soup_one.find_all("a",attrs={"class":"selected"})[2].text.strip()
                urlCity =link.split("//")[1].split(".")[0]
                whetherOther = (urlCity!=city)


                for i in range(1,numOfPage+1):
                    linkDetail = link+ "pg"+str(i)
                    responseDetail =  requests.get(linkDetail)
                    soup_detail =  BeautifulSoup(responseDetail.text,"lxml")
                    soupDetailList  = soup_detail.find("ul",attrs={"class":"sellListContent"}).find_all("li")
                    for li in soupDetailList:
                        pageLink = li.find("div",attrs={"class":"title"}).find("a")["href"]
                        if not pageLink in pageLinks:
                            pageLinks.append(pageLink)
                            title = None
                            title = li.find("a",attrs={"class":"no_resblock_a"}).text.strip()
                            if not title in self.titleList:
                                self.titleList.append(title)
                                #记得补充titleList
                                # print(title)
                                typeOfHouse = li.find("div",attrs={"class":"positionInfo"}).text.split(")")[1].split("-")[0].strip()
                                price = li.find("div",attrs={"class":"unitPrice"})["data-price"]
                                buildYear = "未获取"
                                row = [regionOne,regionTwo,regionThree,others,title,typeOfHouse,buildYear,price]
                                print(row)
                                self.write_sheet.append(row)
            except Exception as e:
                count = count+1
                if count == 5:
                    with open('./doc/error/error.txt',"a+",encoding="utf-8") as ferr:
                        ferr.write("出错的链接是"+link+"，错误的原因是"+str(e)+"\n")
                    with open('./doc/error/errLink.txt',"a+",encoding="utf-8") as ferrLink:
                        ferrLink.write(link+"\n")
            else:
                break

        self.wb.save("./doc/data2.xlsx")
        print("正在保存文件")
    '''
        crawl all the pages in the cityList
    '''
    def getDetail(self):
        print("开始加载excel")
        self.wb = load_workbook("./doc/data2.xlsx")
        self.write_sheet = self.wb['Sheet1']
        for city in self.cityList:
            self.titleList = []
            self.ensureLinks(city)
            if city in self.linkDict.keys():
                linkList = self.linkDict[city]
                for link in linkList:
                    self.getDetailPage(link,city)
    def getNewLinks(self):
        cityList =self.cityList
        for city in cityList:
            linkList = []
            link_all = "https://"+city+".fang.lianjia.com"
            link_one = "https://"+city+".fang.lianjia.com/loupan/"
            response_one = requests.get(link_one,verify=False)
            soup_one = BeautifulSoup(response_one.text,"lxml")
            regionTwoListLinks =soup_one.find_all("li",attrs={"class":"district-item"})
            for regionTwoLink in regionTwoListLinks:
                district_spell = regionTwoLink["data-district-spell"]
                link_two = link_one +district_spell+"/#"+district_spell
                if not  link_two in linkList:
                    linkList.append(link_two)
                    with open("./doc/link/xinfang-"+city+".txt","a+",encoding="utf-8") as f:
                        f.write(link_two+"\n")
        self.linkNewDict[city] = linkList
    def ensureNewLinks(self,city):
        if not city in self.linkNewDict.keys() or len(self.linkNewDict[city])<1:
            print(True)
            with open("./doc/link/xinfang-"+city+".txt","r",encoding="utf-8") as f:
                linkList = []
                lines = list(f)
                for line in lines:
                    linkList.append(line.split("\n")[0])
                self.linkNewDict[city] = linkList
                print("ensureLinks:",self.linkNewDict[city])
    def getNewDetailPage(self,link,city):
        cityDict=self.cityDict
        regionOne = self.cityDict[city]
        print(regionOne)
        count = 0
        others = False
        cityFromURL= link.split("//")[1].split(".")[0]
        others = (city!=cityFromURL)
        pageLinks=[]
        while count <5:
            try:
                responseOne = requests.get(link)
                soup_one =    BeautifulSoup(responseOne.text,"lxml")
                value = int(soup_one.find("span",attrs={"class":"value"}).text)
                numOfPage = value // 10 + 1
                print(numOfPage)
                urlCity =link.split("//")[1].split(".")[0]
                whetherOther = (urlCity!=city)

                for i in range(1,numOfPage+1):
                    linkDetail = link.split("#")[0]+ "pg"+str(i)+"/#"+link.split("#")[1]
                    print(linkDetail)
                    responseDetail =  requests.get(linkDetail)
                    soup_detail =  BeautifulSoup(responseDetail.text,"lxml")
                    soupDetailList  = soup_detail.find_all("li",attrs={"class":"resblock-list"})
                    for li in soupDetailList:
                        pageLink = li.find("a",attrs={"class":"name"})["href"]
                        if not pageLink in pageLinks:
                            pageLinks.append(pageLink)
                            title = None
                            title = li.find("a",attrs={"class":"name"}).text.strip()
                            regions =  li.find("div",attrs={"class":"resblock-location"}).find_all("span")
                            regionTwo = regions[0].text.strip()
                            regionThree =  regions[1].text.strip()
                            typeOfHouse = "未获取"
                            buildYear = "未获取"
                            price = li.find("span",attrs={"class":"number"}).text
                            row = [regionOne,regionTwo,regionThree,others,title,typeOfHouse,buildYear,price]
                            print(row)
                            self.write_sheet.append(row)
            except Exception as e:
                count = count+1
                if count == 5:
                    with open('./doc/error/error.txt',"a+",encoding="utf-8") as ferr:
                        ferr.write("出错的链接是"+link+"，错误的原因是"+str(e)+"\n")
                    with open('./doc/error/errLink.txt',"a+",encoding="utf-8") as ferrLink:
                        ferrLink.write(link+"\n")
            else:
                break
        print("正在保存文件")
        self.wb.save("./doc/data2.xlsx")
    def getNewDetail(self):
        print("开始加载excel")
        self.wb = load_workbook("./doc/data2.xlsx")
        self.write_sheet = self.wb['Sheet1']
        for city in self.cityList:
            self.titleList = []
            self.ensureNewLinks(city)
            if city in self.linkNewDict.keys():
                linkList = self.linkNewDict[city]
                for link in linkList:
                    self.getNewDetailPage(link,city)
